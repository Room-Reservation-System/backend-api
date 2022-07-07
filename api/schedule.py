from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from time import strftime, gmtime
from datetime import datetime, timedelta
import os


class Base():
    wb=Workbook()
    sheet=wb.active
    def __init__(self):
        self.max_days=len(classes.keys())
        self._week_list=['Monday',
                         'Tuesday', 
                         'Wednesday', 
                         'Thursday', 
                         'Friday', 
                         'Saturday', 
                         'Sunday']
        self._xlxs_columns=[['A'],  
                            ['B','C'],
                            ['D','E'],
                            ['F','G'],
                            ['H','I'], 
                            ['J','K'],
                            ['L','M'],
                            ['N','O'],
                            ]

class Template(Base):
  
    def __init__(self,workDays:int=5, startDay:str='Monday'):
        self.workDays=workDays
        self.startDay=startDay
    
    def data(self):
        return self.workDays

demo=Template(workDays=6, startDay='Friday').data()
print(demo)
# class Schedule(Creator):

#     def create_template(self):
#         loc_sheet=Schedule().sheet
#         for col in self._xlxs_columns:
#             loc_sheet.column_dimensions[col[0]].width=25
#             loc_sheet.column_dimensions[col[-1]].width=25
    
#     def create_schedule(self, title:str, classes:dict):


#         for column in self._xlxs_columns:
#             Schedule().sheet.column_dimensions[column[0]].width=25
#             Schedule().sheet.column_dimensions[column[-1]].width=25

        
        # Schedule().sheet.merge_cells('A1:L1')
        # Schedule().sheet['A1']=title
        # Schedule().sheet['A1'].alignment=Alignment(horizontal='center')
        # Schedule().sheet['A1'].font=Font(b=True,size=18)

        
        # sheet.merge_cells('A3:A4')
        # sheet['A3']='Duration'
        # sheet['A3'].alignment=Alignment(horizontal='center',
        #                                 vertical='center')
        # sheet['A3'].font=Font(b=True,size=14)
        # sheet['A3'].fill=PatternFill(start_color="E0E0E0",
        #                              end_color="E0E0E0",
        #                              fill_type="solid")
        # my_time=[timedelta(hours=9, minutes=0)]
        # for i in range(44):
        #     my_time.append((my_time[-1]+timedelta(minutes=15)))
        # for row in my_time:
        #     print(':'.join(str(row).split(':')[:2]))

        # for i in range(1,len(self.classes.keys())+1):
        #     sheet.merge_cells(f'{self._xlxs_columns[i][0]}3:{self._xlxs_columns[i][1]}3')
        #     sheet[f'{self._xlxs_columns[i][0]}3']=self._week_list[i-1]
        #     sheet[f'{self._xlxs_columns[i][0]}3'].alignment=Alignment(horizontal='center',
        #                                                               vertical='center')
        #     sheet[f'{self._xlxs_columns[i][0]}3'].font=Font(b=True,size=16)
        #     sheet[f'{self._xlxs_columns[i][0]}3'].fill=PatternFill(start_color="E0E0E0",
        #                                                              end_color="E0E0E0",
        #                                                              fill_type="solid")


        # Schedule.wb.save(filename=os.path.join(self._base_dir,"test1.xlsx"))
        # Schedule.wb.save('demo.xlsx')
# # print(strftime('%A', gmtime()))
# title='Timetable of Spring period (January - May) of 2021-2022 Academic Years for Undergraduate Year 1 Program.'
# classes={'Monday':{},
#          'Tuesday':{}, 
#          'Wednesday':{}, 
#          'Thursday':{}, 
#          'Friday':{}, 
#         #  'Saturday':{}, 
#         #  'Sunday':{}
#          }

# demo=Schedule()
# demo.create_schedule(title=title,classes=classes)
# print(settings.EMAIL_HOST_USER)

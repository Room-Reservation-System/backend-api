from random import choice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from datetime import datetime
from .filter import Filter


class Base:
    wb=Workbook()
    sheet=wb.active

location=Alignment(horizontal='center',vertical='center',)

font_main_bold=Font(name='Arial',size=16,bold=True,color='FF000000',)

font_main=Font(name='Arial',size=16,bold=False,color='FF000000',)

font_title=Font(name='Arial',size=22,bold=True,color='FF000000',)

font_subject=Font(name='Arial',size=16,bold=False,color='FF000000',)

font_event=Font(name='Arial', size=16,bold=False,color='FF000000',)

fill_main=PatternFill(fill_type='solid',start_color='E0E0E0',end_color='E0E0E0',)   
# BORDER_NONE = None
# BORDER_DASHDOT = 'dashDot'
# BORDER_DASHDOTDOT = 'dashDotDot'
# BORDER_DASHED = 'dashed'
# BORDER_DOTTED = 'dotted'
# BORDER_DOUBLE = 'double'
# BORDER_HAIR = 'hair'
# BORDER_MEDIUM = 'medium'
# BORDER_MEDIUMDASHDOT = 'mediumDashDot'
# BORDER_MEDIUMDASHDOTDOT = 'mediumDashDotDot'
# BORDER_MEDIUMDASHED = 'mediumDashed'
# BORDER_SLANTDASHDOT = 'slantDashDot'
# BORDER_THICK = 'thick'
# BORDER_THIN = 'thin'

border_medium=Border(left=Side(border_style='medium', color='000000'),
                     right=Side(border_style='medium',color='000000'),
                     top=Side(border_style='medium',color='000000'),
                     bottom=Side(border_style='medium', color='000000'),)

border_group=Border(left=Side(border_style='medium', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='medium',color='000000'),)

def templateText(column:str,row:int,value:str,name:str='Arial',size:int=16,bold:bool=False,color='FF000000'):
    Base.sheet[f'{column}{row}']=value
    Base.sheet[f'{column}{row}'].font=Font(name=name,size=size,bold=bold,color=color,)
    Base.sheet[f'{column}{row}'].alignment=Alignment(horizontal='center', vertical='center',)

def templateDescription(column:str,row:int,value:str,name:str='Arial',size:int=14,bold:bool=False,color='FF000000'):
    Base.sheet[f'{column}{row}']=value
    Base.sheet[f'{column}{row}'].font=Font(name=name,size=size,bold=bold,color=color,)
    Base.sheet[f'{column}{row}'].alignment=Alignment(wrap_text=True, horizontal='center',vertical='center',)

def colorCell(column:str,row:int,color:str='#FFFFFF',):
    Base.sheet[f'{column}{row}'].fill=PatternFill(fill_type='solid',start_color=color,end_color=color, )

class Node():
    current_index:int=0
    used:int=0
    next_index:int=0

class Base:
    wb=Workbook()
    sheet=wb.active

class TableGenerator(Base):
    def __init__(self,data:list,title:str,timing:dict={'startTime': {'hours':8, 'minutes':00},'endTime':{'hours':24,'minutes':00}},):

        self.data=Filter().filter(data)
        self.title=f"Time table of {title}"
        self.timing=timing

        self.week_list=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
        self.columns=['A','B','C','D','E','F','G','H']

    def setData(self,):
        color_fill=[ #'FFCCCC','FFE5CC','FFFFCC','E5FFCC','CCFFCC','CCFFE5','CCFFFF','CCE5FF','CCCCFF','E5CCFF','FFCCFF','FFCCE5',
                    'FF9999','FFCC99','FFFF99','CCFF99','99FF99','99FFCC','99FFFF','99CCFF','9999FF','CC99FF','FF99FF','FF99CC'
        ]

        self.__getTemplate()

        subjectColor:dict={}
       
        for val in self.data:
            
            row_A:int=0
            row_B:int=0        
            column_A:int=''
            name=val['title'].upper()
            start_time=val['start_time']
            end_time=val['end_time']
            if name not in subjectColor:
                subjectColor[name]=choice(color_fill)
                print(102)
            
            for i, row in enumerate(Base.sheet['A']):
                if row.value==val['start_time']:
                    row_A=i
                if row.value==val['end_time']:
                    row_B=i
            for i, column in enumerate(Base.sheet[3]):
                if column.value==val['day']:
                    column_A=i
            cellDesc:str=f'{name}\n\n{start_time}-{end_time}'
            colorCell(column=self.columns[column_A],row=row_A+1, color=subjectColor[name])
            Base.sheet.merge_cells(f'{self.columns[column_A]}{row_A+1}:{self.columns[column_A]}{row_B}')
            templateDescription(column=self.columns[column_A],row=row_A+1,value=cellDesc,size=14)

        for key, val in subjectColor.items():
            print(f'{key}: {val}')
        Base.wb.save(filename='test.xlsx')

    def __getTemplate(self):
        Base.sheet.merge_cells('B1:H1')
        Base.sheet['B1'].fill=fill_main
        Base.sheet['A1'].fill=fill_main
        templateText(column='B', row=1, size=20, value=self.title)
        for i, cell in enumerate(self.columns):
            Base.sheet.merge_cells(f'{self.columns[i]}3:{self.columns[i]}4')
            Base.sheet[f'{self.columns[i]}3'].fill=fill_main
            if cell=='A':
                Base.sheet.column_dimensions['A'].width=15
                templateText(column='A',row=3,value='Duration')
            else:
                Base.sheet.column_dimensions[cell].width=35
                
                templateText(column=cell, row=3, value=self.week_list[i-1])
        data_time=[hour%24 for hour in range (self.timing['startTime']['hours'],self.timing['endTime']['hours']+1)]
        start_minute=self.timing['startTime']['minutes']
        minutes=[]
        for hour in data_time:
            row=[]
            for minute in range ((60-start_minute)//30):
                row.append(start_minute)
                start_minute+=30
            minutes.append(row)
            start_minute=0
            if hour==self.timing['endTime']['hours'] and start_minute==self.timing['endTime']['minutes']:
                minutes[-1].append(start_minute)
                break
        for i, value in enumerate(data_time):
            for minute in minutes[i]:
                row=self.__getIndex(initial_index=6)
                time_var:str=f'{value}:{minute}'
                if time_var[-2:]==':0':time_var+='0'
                
                if self.timing['endTime']['hours']%24==value and self.timing['endTime']['minutes']<minute:
                    m=str(self.timing['endTime']['minutes'])
                    if m=='0':m+='0'
                    h=data_time[-1]
                    break 
                templateText(column='A', row=row, value=time_var,size=14)
                Base.sheet.row_dimensions[row].height=30
                Base.sheet[f'A{row}'].fill=fill_main
        templateText(column='A', row=5, value=f'{data_time[0]}:{minutes[0][0]}-{h}:{m}', bold=True)
        
        self.__clearNode()

    def __getIndex(self,initial_index:int=0):
        local_index:int=Node.current_index+initial_index
        Node.current_index+=1
        return local_index

    def __clearNode(self,):
        Node.current_index=0
        Node.used=0
        Node.next_index=0



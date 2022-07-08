
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from time import strftime, gmtime, time
from datetime import datetime, timedelta
import os
from base import Base 

class Node():
    current_index:int=0
    used:int=0

class Template(Base, Node):


    
    def create_template(self, 
                        title:str, 
                        workDays:int=7, 
                        startWork:dict={'hours':9, 
                                        'minutes':00},
                        endWork:dict={'hours':20, 
                                      'minutes':00},
                        startDay:str='Monday',
                        cohort:dict={'cohort1':'Cohort One: Arts',
                                     'cohort2':'Cohort Two: Sciences'},
                        # cohort=None,
                        ):

        sheet=Base.sheet
        locate=Base.location

        for cell in range(workDays*2+1):
            sheet.column_dimensions[self._columns[cell]].width=25

        
        sheet.merge_cells(f'B1:{self._columns[workDays*2]}1')
        sheet['B1'].alignment=locate
        sheet['B1'].font=Font(b=True,size=18)
        sheet['B1']=title

        for cell in range(1,workDays*2+1,2):
            if cohort is None:
                sheet.merge_cells(f'{self._columns[cell]}3:{self._columns[cell+1]}4')
                sheet[f'{self._columns[cell]}3'].alignment=locate
                sheet[f'{self._columns[cell]}3'].font=Font(b=True,size=16)
                sheet[f'{self._columns[cell]}3']=self.__getWeek(workDays=workDays)
            else:
                sheet.merge_cells(f'{self._columns[cell]}3:{self._columns[cell+1]}3')
                sheet[f'{self._columns[cell]}4']=cohort['cohort1']
                sheet[f'{self._columns[cell]}4'].alignment=locate
                sheet[f'{self._columns[cell]}4'].font=Font(b=False,size=12)

                sheet[f'{self._columns[cell+1]}4']=cohort['cohort2']
                sheet[f'{self._columns[cell+1]}4'].alignment=locate
                sheet[f'{self._columns[cell+1]}4'].font=Font(b=False,size=12)
                
                sheet[f'{self._columns[cell]}3'].alignment=locate
                sheet[f'{self._columns[cell]}3'].font=Font(b=True,size=16)
                sheet[f'{self._columns[cell]}3']=self.__getWeek(workDays=workDays)
                
    # def __getCell(self, cell_id=0):
    #     Node.current_index=cell_id
    #     local_id=Node.current_index
    #     Node.current_index+=1
    #     Node.used=cell_id
    #     if Node.used==cell_id:
    #         Node.used+=1
    #         return Node.used
    #     return local_id

    def __getWeek(self, workDays:int=5):
        loc_current=Node.current_index
        Node.current_index+=1
        if Node.current_index==workDays:Node.current_index=0
        return self._week_list[loc_current]

    def add_week(self):
        local_sheet=Base.sheet
        local_sheet.merge_cells('A2:B2')
        local_sheet['A2']='hello'

    def save_template(self,filename:str):
        Base.wb.save(filename=f'{filename}.xlsx')

title='Timetable of Spring period (January - May) of 2021-2022 Academic Years for Undergraduate Year 1 Program.'
template=Template()
template.create_template(title=title, workDays=5)
template.save_template('demo2')
# template.add_week()
# template.save_template('demo1')


# demo=Template(workDays=6, startDay='Friday').data()
# demo=Template().data()
# print(demo)
# print(demo)
# class Schedule(Creator):

#     def create_template(self):
#         loc_sheet=Schedule().sheet
#         for col in self._xlxs_columns:
#             loc_sheet.column_dimensions[col[0]].width=25
#             loc_sheet.column_dimensions[col[-1]].width=25
    
#     def create_schedule(self, title:str, classes:dict):


#         for column in self._xlxs_columns:
#             Schedule().sheet.column_dimensions[column[0]].width=25
#             Schedule().sheet.column_dimensions[column[-1]].width=25

        
        # Schedule().sheet.merge_cells('A1:L1')
        # Schedule().sheet['A1']=title
        # Schedule().sheet['A1'].alignment=Alignment(horizontal='center')
        # Schedule().sheet['A1'].font=Font(b=True,size=18)

        
        # sheet.merge_cells('A3:A4')
        # sheet['A3']='Duration'
        # sheet['A3'].alignment=Alignment(horizontal='center',
        #                                 vertical='center')
        # sheet['A3'].font=Font(b=True,size=14)
        # sheet['A3'].fill=PatternFill(start_color="E0E0E0",
        #                              end_color="E0E0E0",
        #                              fill_type="solid")
        # my_time=[timedelta(hours=9, minutes=0)]
        # for i in range(44):
        #     my_time.append((my_time[-1]+timedelta(minutes=15)))
        # for row in my_time:
        #     print(':'.join(str(row).split(':')[:2]))

        # for i in range(1,len(self.classes.keys())+1):
        #     sheet.merge_cells(f'{self._xlxs_columns[i][0]}3:{self._xlxs_columns[i][1]}3')
        #     sheet[f'{self._xlxs_columns[i][0]}3']=self._week_list[i-1]
        #     sheet[f'{self._xlxs_columns[i][0]}3'].alignment=Alignment(horizontal='center',
        #                                                               vertical='center')
        #     sheet[f'{self._xlxs_columns[i][0]}3'].font=Font(b=True,size=16)
        #     sheet[f'{self._xlxs_columns[i][0]}3'].fill=PatternFill(start_color="E0E0E0",
        #                                                              end_color="E0E0E0",
        #                                                              fill_type="solid")


        # Schedule.wb.save(filename=os.path.join(self._base_dir,"test1.xlsx"))
        # Schedule.wb.save('demo.xlsx')
# # print(strftime('%A', gmtime()))
# title='Timetable of Spring period (January - May) of 2021-2022 Academic Years for Undergraduate Year 1 Program.'
# classes={'Monday':{},
#          'Tuesday':{}, 
#          'Wednesday':{}, 
#          'Thursday':{}, 
#          'Friday':{}, 
#         #  'Saturday':{}, 
#         #  'Sunday':{}
#          }

# demo=Schedule()
# demo.create_schedule(title=title,classes=classes)
# print(settings.EMAIL_HOST_USER)

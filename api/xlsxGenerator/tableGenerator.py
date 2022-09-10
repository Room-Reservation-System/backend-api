from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font)
from random import choice
from .base import Base
from .filter import Filter
from .objectsType import ScheduleTime, Group, TypeXlsx
from .cleaner import clearAll
from django.conf import settings
from os import path


class Node():
    current_index:int=0
    used:int=0
    next_index:int=0


class TableGenerator:

    def __init__(self,title:str,workHours:ScheduleTime=ScheduleTime(startHour=9, startMinute=0, endHour=24, endMinute=0),step:int=15,):

        self.title=f'Timetable for {title}'
        self.workHours=workHours
        self.step=step
        self.fileName=f'{self.title}.xlsx'
        self.dirName=path.join(settings.BASE_DIR, 'xlsxFiles')
        self.week_list=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
        self.base=Base(self.fileName)
        self.sheet=self.base.getWorkSheet()
        self.roomColor:dict={'201':'FF0000',
                        '202':'FFFFFF',
                        '203':'D0CECE',
                        '204':'92D050',
                        '206':'CC99FF',
                        '209':'FFFF00',
                        '109':'F8CBAD',
                        '111':'F4B084',
                        'B14':'9BC2E6',
                        '3.28':'D6DCE4',
                        'B14':'8EA9DB',
                        '11':'00FFFF',
                        '69':'FFFFFF'}
        self.setTimeColumn()

    def setDataCohortMode(self,data,group:Group=Group(groupOne='Cohort One: Arts',groupTwo='Cohort Two: Sciences')):
        groupOne,groupTwo=group

        columns=['A','B','C','D','E','F','G','H','I','J','K']
        self.setBaseTemplate(start='B',end=columns[-1])
        self.getCohortTemplate(group=group)
        for val in data:

            row_A:int=0
            row_B:int=0
            columnCS:int=0
            name=val['title'].upper()
            start_time=val['start_time']
            end_time=val['end_time']
            room=val['room']

            for i, row in enumerate(self.sheet['A']):
                if row.value==val['start_time']:
                    row_A=i #1
                if row.value==val['end_time']:
                    row_B=i#2
            for i in range(1,len(columns)):
                if self.sheet[f'{columns[i]}{3}'].value==val['day']:
                    columnCS=i
                    if val['group']=='CS':
                        columnCS+=1

            cellDesc:str=f'{name}\n\n{start_time}-{end_time}\nroom:{room}'
            self.borderSet(column=columnCS+1, startRow=row_A+1, endRow=row_B)
            try:
                self.colorCell(column=columns[columnCS],row=row_A+1, color=self.roomColor[str(val['room'])])
            except:
                self.colorCell(column=columns[columnCS],row=row_A+1, color=self.roomColor["202"])
            self.sheet.merge_cells(f'{columns[columnCS]}{row_A+1}:{columns[columnCS]}{row_B}')
            self.writeText(column=columns[columnCS],row=row_A+1,text=cellDesc, fontType='class')

        self.base.saveXlsx()

    def setDataRoomMode(self,data):
        columns=['A','B','C','D','E','F','G','H']

        self.setBaseTemplate(start='B',end=columns[-1])
        self.getRoomTemplate()

        for val in data:

            row_A:int=0
            row_B:int=0
            column_A:int=''
            instructor=val['instructor']
            name=val['title'].upper()
            start_time=val['start_time']
            end_time=val['end_time']

            for i, row in enumerate(self.sheet['A']):
                if row.value==val['start_time']:
                    row_A=i #1
                if row.value==val['end_time']:
                    row_B=i#2
            for i, column in enumerate(self.sheet[3]):
                if column.value==val['day']:
                    column_A=i
            cellDesc:str=f'{name}\n\n{start_time}-{end_time}'
            self.borderSet(column=column_A+1, startRow=row_A+1, endRow=row_B)

            try:
                self.colorCell(column=columns[column_A],row=row_A+1, color=self.roomColor[str(val['room'])])
            except:
                self.colorCell(column=columns[column_A],row=row_A+1, color=self.roomColor["202"])
            self.sheet.merge_cells(f'{columns[column_A]}{row_A+1}:{columns[column_A]}{row_B}')
            self.writeText(column=columns[column_A],row=row_A+1,text=cellDesc, fontType='class')

        self.base.saveXlsx()

    def setDataFaculty(self,data):
        columns=['A','B','C','D','E','F','G','H']

        self.setBaseTemplate(start='B',end=columns[-1])
        self.getRoomTemplate()
        for val in data:

            row_A:int=0
            row_B:int=0
            column_A:int=''
            name=val['title'].upper()
            start_time=val['start_time']
            end_time=val['end_time']

            for i, row in enumerate(self.sheet['A']):
                if row.value==val['start_time']:
                    row_A=i #1
                if row.value==val['end_time']:
                    row_B=i#2
            for i, column in enumerate(self.sheet[3]):
                if column.value==val['day']:
                    column_A=i
            cellDesc:str=f'{name}\n\n{start_time}-{end_time}'
            self.borderSet(column=column_A+1, startRow=row_A+1, endRow=row_B)
            try:
                self.colorCell(column=columns[column_A],row=row_A+1, color=self.roomColor[str(val['room'])])
            except:
                self.colorCell(column=columns[column_A],row=row_A+1, color=self.roomColor["202"])
            self.sheet.merge_cells(f'{columns[column_A]}{row_A+1}:{columns[column_A]}{row_B}')
            self.writeText(column=columns[column_A],row=row_A+1,text=cellDesc, fontType='class')

        self.base.saveXlsx()

    def getCohortTemplate(self,group:Group):

        columns=['A','B','C','D','E','F','G','H','I','J','K']
        #resize cells and coloring:
        # for cell in (1,66):
        #     self.sheet.row_dimensions[cell].height=10
        for col in range(len(columns)):
            self.colorCell(column=columns[col],row=3)
            if columns[col] == 'A':
                self.sheet.column_dimensions[columns[col]].width=15
            else:
                self.sheet.column_dimensions[columns[col]].width=35
        # weeksetter
        for cellid in range(1,len(columns),2):
            self.sheet.merge_cells(f'{columns[cellid]}3:{columns[cellid+1]}3')
            self.writeText(column=columns[cellid], row=3, text=self.week_list[self.getIndex(0)],fontType='general')
        self.clearNode()
        #naming groups
        for ind, val in enumerate(columns[1:]):
            self.colorCell(column=val, row=4)
            if ind%2:
                self.writeText(column=val, row=4,text=group.groupTwo,fontType='general')
            else:
                self.writeText(column=val, row=4,text=group.groupOne,fontType='general')

        # duration
        self.sheet.merge_cells('A3:A4')
        self.writeText(column='A',row=3,text='Dutaion')

        self.base.saveXlsx()

    def getRoomTemplate(self):
        columns=['A','B','C','D','E','F','G','H']


        for i, cell in enumerate(columns):
            self.sheet.merge_cells(f'{columns[i]}3:{columns[i]}4')
            self.colorCell(column=columns[i],row=3)
            if cell=='A':
                self.sheet.column_dimensions['A'].width=15
                self.writeText(column='A',row=3,text='Duration')
            else:
                self.sheet.column_dimensions[cell].width=35
                self.writeText(column=cell, row=3, text=self.week_list[i-1],fontType='general')

    def textLocation(self,column:str,row:int,wrap_text=True):
        self.sheet[f'{column}{row}'].alignment=Alignment(horizontal='center',vertical='center',wrap_text=wrap_text)

    def fontText(self,column:str,row:int,fontType:str):
        locStyles:dict={'general':{'name':'Arial','size':16,'bold':False,'color':'00000000'},
                        'generalBold':{'name':'Arial','size':16,'bold':True,'color':'00000000'},
                        'title':{'name':'Arial','size':22,'bold':True,'color':'00000000'},
                        'event':{'name':'Arial','size':16,'bold':False,'color':'00000000'},
                        'class':{'name':'Arial','size':14,'bold':True,'color':'00000000'},
                        'personal':{},}
        self.sheet[f'{column}{row}'].font=Font(**locStyles[fontType])

    def colorCell(self,column:str,row:int,color=None):
        if color is None:
            color='E0E0E0'
        self.sheet[f'{column}{row}'].fill=PatternFill(fill_type='solid',start_color=color,end_color=color, )

    def borderStyle(self,column:str,row:int,borderType:str):
        locBorder=Side(border_style=borderType)
        self.sheet[f'{column}{row}'].border=Border(left=locBorder,right=locBorder,bottom=locBorder,top=locBorder)

    def writeText(self,column:str,row:int,text:str,fontType:str='general',wrapText:bool=True):
        self.sheet[f'{column}{row}']=text
        self.textLocation(column=column,row=row,wrap_text=wrapText)
        self.fontText(column=column,row=row,fontType=fontType)

    def borderSet(self, column:int, startRow:int, endRow:int):
        locBorder=Side(border_style='medium')
        for row in self.sheet.iter_rows(min_row=startRow,min_col=column,max_row=endRow,max_col=column):
            for cell in row:
                cell.border=Border(left=locBorder,right=locBorder,bottom=locBorder,top=locBorder)

    def getFile(self):
        return path.join(self.dirName, self.fileName)

    def getDir(self):
        return self.dirName

    def getIndex(self,initial_index:int=0):
        local_index:int=Node.current_index+initial_index
        Node.current_index+=1
        return local_index

    def clearNode(self,):
        Node.current_index=0
        Node.used=0
        Node.next_index=0

    def setBaseTemplate(self,start:str,end:str):
        self.sheet.merge_cells(f'{start}1:{end}1')
        self.sheet.row_dimensions[1].height=25
        self.colorCell(column='B',row=1,color='E0E0E0')
        self.colorCell(column='A',row=1,color='E0E0E0')
        self.writeText(column='B', row=1, text=self.title,fontType='title')

    def setTimeColumn(self):
        for cell in (6,67):
            self.sheet.row_dimensions[cell].height=5
        data_time=[hour%24 for hour in range (self.workHours.startHour,self.workHours.endHour+1)]
        start_minute=self.workHours.startMinute
        minutes=[]
        for hour in data_time:
            row=[]
            for minute in range ((60-start_minute)//self.step):
                row.append(start_minute)
                start_minute+=self.step
            minutes.append(row)
            start_minute=0
            if hour==self.workHours.endHour and start_minute==self.workHours.endMinute:
                minutes[-1].append(start_minute)
                break
        for i, value in enumerate(data_time):
            for minute in minutes[i]:
                row=self.getIndex(initial_index=6)
                time_var:str=f'{value}:{minute}'
                if time_var[-2:]==':0':time_var+='0'

                if self.workHours.endHour%24==value and self.workHours.endMinute<minute:
                    m=f'{self.workHours.endMinute}'
                    if m=='0':m+='0'
                    break
                self.writeText(column='A', row=row, text=time_var,fontType='general')
                self.colorCell(column='A', row=row)
                self.sheet.row_dimensions[row].height=30
        self.clearNode()
from openpyxl import Workbook, load_workbook
from django.conf import settings
from os import path, walk, mkdir
from .cleaner import clearAll

class Base:
    
    def __init__(self,filename:str):
        self.fileName=filename
        self.filePath=path.join(settings.BASE_DIR,'xlsxFiles',filename)
        self.folderDir=path.join(settings.BASE_DIR,'xlsxFiles')
        self.workBook=None

    def getWorkSheet(self):
        if not path.isdir(self.folderDir):
            mkdir(self.folderDir)
        files=[]
        for (dPath,dNames,dFiles) in walk(self.folderDir):
            files.extend(dFiles)

        if self.fileName in files:
            with open(self.filePath, 'rb') as xlsx:
                self.workBook=load_workbook(xlsx)
                workSheet=self.workBook[self.workBook.sheetnames[0]]
                return workSheet
        else:
            self.workBook=Workbook()
            workSheet=self.workBook.active
            return workSheet
    
    def saveXlsx(self):
        self.workBook.save(filename=self.filePath)
        self.workBook.close()

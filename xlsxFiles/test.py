from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side


path='C:\\Users\\ekuul\\Desktop\\GreatProjects\\Room reservation system\\main\\backend-api\\xlsxFiles\\Timetable for 203.xlsx'
with open(path, 'rb') as xlsx:
    wb=load_workbook(xlsx)
    ws=wb[wb.sheetnames[0]]
# for row in ws.iter_rows('B{}:B{}'.format(10,12)):
locBorder=Side(border_style='medium')
for row in ws.iter_rows(min_row=14,min_col=2,max_row=19,max_col=2):
    for cell in row:
        cell.border=Border(left=locBorder,right=locBorder,bottom=locBorder,top=locBorder)
wb.save(filename='qweqwe.xlsx')